"""
Excel Writer Module
Menulis data bandwidth ke file Excel yang sudah ada
Dengan fitur skip baris yang sudah terisi
"""

from datetime import datetime
from typing import Dict, List, Optional, Callable
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import config


class ExcelWriter:
    """Writer untuk mengisi data ke Excel existing"""
    
    def __init__(self, file_path: str, progress_callback: Optional[Callable] = None):
        """
        Initialize writer
        
        Args:
            file_path: Path ke file Excel
            progress_callback: Fungsi callback untuk update progress
        """
        self.file_path = file_path
        self.workbook = None
        self.progress_callback = progress_callback or (lambda msg, pct: None)
    
    def _update_progress(self, message: str, percentage: int = -1):
        """Update progress via callback"""
        self.progress_callback(message, percentage)
    
    def open_workbook(self):
        """Buka file Excel"""
        self._update_progress(f"Membuka file Excel...", 86)
        self.workbook = load_workbook(self.file_path)
        self._update_progress(f"File Excel dibuka: {len(self.workbook.sheetnames)} sheet ditemukan")
    
    def save_workbook(self):
        """Simpan file Excel"""
        self._update_progress("Menyimpan file Excel...", 98)
        self.workbook.save(self.file_path)
        self._update_progress("File Excel tersimpan!", 100)
    
    def close_workbook(self):
        """Tutup workbook"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
    
    def get_sheet(self, sheet_name: str) -> Optional[Worksheet]:
        """
        Ambil sheet berdasarkan nama
        
        Args:
            sheet_name: Nama sheet
            
        Returns:
            Worksheet object atau None jika tidak ditemukan
        """
        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        
        # Coba case-insensitive match
        for name in self.workbook.sheetnames:
            if name.lower() == sheet_name.lower():
                return self.workbook[name]
        
        return None
    
    def is_row_filled(self, sheet: Worksheet, row: int) -> bool:
        """
        Cek apakah baris sudah terisi data
        
        Args:
            sheet: Worksheet object
            row: Nomor baris
            
        Returns:
            True jika sudah ada data di kolom Curent(IN)
        """
        cell_value = sheet.cell(row=row, column=config.EXCEL_COL_CURR_IN).value
        return cell_value is not None and str(cell_value).strip() != ""
    
    def find_row_by_date_time(self, sheet: Worksheet, target_date: datetime, 
                               target_hour: int, target_minute: int) -> Optional[int]:
        """
        Cari baris yang sesuai dengan tanggal dan waktu
        
        Args:
            sheet: Worksheet object
            target_date: Tanggal yang dicari
            target_hour: Jam yang dicari
            target_minute: Menit yang dicari
            
        Returns:
            Nomor baris (1-indexed) atau None jika tidak ditemukan
        """
        target_date_str = target_date.strftime(config.DATE_FORMAT_EXCEL)
        
        # Format waktu sesuai config (pakai titik: "09.00")
        target_time_compare = f"{target_hour:02d}.{target_minute:02d}"
        if config.TIME_FORMAT_EXCEL:
            target_time_compare = datetime(2000, 1, 1, target_hour, target_minute).strftime(config.TIME_FORMAT_EXCEL)
        
        for row in range(config.EXCEL_DATA_START_ROW, sheet.max_row + 1):
            date_cell = sheet.cell(row=row, column=config.EXCEL_COL_TANGGAL).value
            time_cell = sheet.cell(row=row, column=config.EXCEL_COL_WAKTU).value
            
            # Konversi nilai cell ke string untuk perbandingan
            date_str = self._cell_to_date_string(date_cell)
            time_str = self._cell_to_time_string(time_cell)
            
            # Bandingkan
            if date_str == target_date_str and time_str == target_time_compare:
                return row
        
        return None
    
    def _cell_to_date_string(self, cell_value) -> str:
        """Konversi nilai cell tanggal ke string"""
        if cell_value is None:
            return ""
        if isinstance(cell_value, datetime):
            return cell_value.strftime(config.DATE_FORMAT_EXCEL)
        return str(cell_value).strip()
    
    def _cell_to_time_string(self, cell_value) -> str:
        """Konversi nilai cell waktu ke string"""
        if cell_value is None:
            return ""
        if isinstance(cell_value, datetime):
            return cell_value.strftime(config.TIME_FORMAT_EXCEL)
        # Handle time object
        if hasattr(cell_value, 'strftime'):
            return cell_value.strftime(config.TIME_FORMAT_EXCEL)
        # Handle numeric time (e.g., 9.0 for 09.00)
        if isinstance(cell_value, (int, float)):
            hours = int(cell_value)
            minutes = int((cell_value - hours) * 100)
            return f"{hours:02d}.{minutes:02d}"
        return str(cell_value).strip()
    
    def write_data_to_row(self, sheet: Worksheet, row: int, data: Dict):
        """
        Tulis data bandwidth ke baris tertentu
        
        Args:
            sheet: Worksheet object
            row: Nomor baris
            data: Dictionary dengan data bandwidth
        """
        # Mapping data ke kolom
        column_mapping = {
            config.EXCEL_COL_CURR_IN: data.get('curr_in'),
            config.EXCEL_COL_CURR_OUT: data.get('curr_out'),
            config.EXCEL_COL_MAX_IN: data.get('max_in'),
            config.EXCEL_COL_MAX_OUT: data.get('max_out'),
            config.EXCEL_COL_AVG_IN: data.get('avg_in'),
            config.EXCEL_COL_AVG_OUT: data.get('avg_out'),
        }
        
        for col, value in column_mapping.items():
            if value is not None:
                sheet.cell(row=row, column=col, value=value)
    
    def write_all_data(self, data_list: List[Dict]):
        """
        Tulis semua data ke Excel
        
        Args:
            data_list: List of data dictionaries from scraper
        """
        total = len(data_list)
        written = 0
        skipped = 0
        not_found = 0
        
        for i, data in enumerate(data_list):
            sheet_name = data.get('sheet')
            if not sheet_name:
                continue
            
            sheet = self.get_sheet(sheet_name)
            if not sheet:
                self._update_progress(f"⚠ Sheet '{sheet_name}' tidak ditemukan!")
                continue
            
            # Cari baris yang sesuai
            row = self.find_row_by_date_time(
                sheet, 
                data['date'],
                data['time_hour'],
                data['time_minute']
            )
            
            if row:
                date_str = data['date'].strftime(config.DATE_FORMAT_EXCEL)
                time_str = f"{data['time_hour']:02d}.{data['time_minute']:02d}"
                
                # Cek apakah baris sudah terisi
                if config.SKIP_FILLED_ROWS and self.is_row_filled(sheet, row):
                    skipped += 1
                    self._update_progress(
                        f"⏭ Skip {sheet_name}: {date_str} {time_str} (sudah terisi)",
                        86 + int((i / total) * 12)
                    )
                else:
                    self.write_data_to_row(sheet, row, data)
                    written += 1
                    self._update_progress(
                        f"✓ {sheet_name}: {date_str} {time_str} → baris {row}",
                        86 + int((i / total) * 12)
                    )
            else:
                not_found += 1
                date_str = data['date'].strftime(config.DATE_FORMAT_EXCEL)
                time_str = f"{data['time_hour']:02d}.{data['time_minute']:02d}"
                self._update_progress(f"✗ Tidak ditemukan: {sheet_name} {date_str} {time_str}")
        
        self._update_progress(f"Total: {written} ditulis, {skipped} di-skip, {not_found} tidak ditemukan")


def write_to_excel(file_path: str, data_list: List[Dict],
                   progress_callback: Optional[Callable] = None):
    """
    Fungsi utama untuk menulis data ke Excel
    
    Args:
        file_path: Path ke file Excel
        data_list: List data dari scraper
        progress_callback: Callback untuk progress update
    """
    writer = ExcelWriter(file_path, progress_callback)
    
    try:
        writer.open_workbook()
        writer.write_all_data(data_list)
        writer.save_workbook()
    finally:
        writer.close_workbook()
